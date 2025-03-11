from openpyxl.utils import get_column_letter

from snapshot_structures import *
from colors import Color, pattern_fill

import datetime
current_time = datetime.datetime.now()

from re import search

###Caseload Details highlights
# Highlight No values
def highlight_no_values(worksheet):
    cm_status_col = cld_col_num['Case Managed Status']
    id_col = cld_col_num['Student ID']
    parent = cld_col_num['Parent Consent + Supports Within Consent Date?']
    sna_col = cld_col_num['Student Needs Assessment Entered?']
    ssp_col = cld_col_num['Student Support Plan Entered?']
    checkin_col = cld_col_num['At least 1 check-in per month?']
    # Check validity of Student ID
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=id_col, max_col=id_col):
        for cell in row:
            m = search('^(\d{6})[MFX](\d{3})', cell.value)
            if m is None:
                pattern_fill(cell, color=Color.RED)
                
    # check for parent consent
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=parent, max_col=parent):
        for cell in row:
            if cell.value == 'No':
                pattern_fill(cell, color=Color.YELLOW)
    # SNA & SSP check
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=sna_col, max_col=ssp_col):
        for cell in row:
            # don't highlight in Jan for students enrolled less than 1 month
            months_enrolled = worksheet['M' + str(cell.row)].value
            if current_time.month == 1 and months_enrolled == 0:
                pass
            else:
                if cell.value == 'No':
                    cm_status = worksheet[get_column_letter(cm_status_col) + str(cell.row)]
                    if cm_status.value == 'Case Managed':
                        pattern_fill(cell, color=Color.YELLOW)
    # 'At least 1 check-in per month?'                
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=checkin_col, max_col=checkin_col):
        for cell in row:
            if cell.value == 'No':
#                 cm_status = worksheet[get_column_letter(cm_status_col) + str(cell.row)]
#                 if cm_status.value == 'Case Managed':
                pattern_fill(cell, color=Color.YELLOW)
    

# Highlight students that should be switched to non-case managed
def highlight_cm_switch(worksheet):
    cm_status_col = cld_col_num['Case Managed Status']
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=cm_status_col, max_col=cm_status_col):
        for cell in row:
            if cell.offset(column=2).value != datetime.datetime(9999, 12, 31, 0, 0):
                if cell.value == "Case Managed":
                    if cell.offset(column=6).value == 'No' or cell.offset(column=7).value == 'No':
                        pattern_fill(cell, color=Color.YELLOW)
                    
    
# Highlight SEL assessments under the minimum in Caseload Details sheet
# set a date for changing number of required sel assessments
second_year = years[1]
sel_date_change = datetime.datetime(second_year, 5, 1)

def highlight_SEL(worksheet):
    sel_num = cld_col_num['SEAD/DESSA']
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=sel_num, max_col=sel_num):
        for cell in row:
            # don't highlight in Jan for students enrolled less than 1 month
            months_enrolled = worksheet['M' + str(cell.row)].value
            if current_time.month == 1 and months_enrolled == 0:
                pass
            # check if student is currently enrolled
            elif cell.offset(column=-3).value == datetime.datetime(9999, 12, 31, 0, 0):
                # highlight for case managed students
                if cell.offset(column=-5).value == 'Case Managed':
                    if today < sel_date_change and cell.value < 1:
                        pattern_fill(cell, color=Color.YELLOW)
                    elif today >= sel_date_change and cell.value < 2:
                        pattern_fill(cell, color=Color.YELLOW)
                # highlight for non-case managed students
                else:
                    if today >= sel_date_change and cell.value < 1:
                        pattern_fill(cell, color=Color.YELLOW)
                    

# Highlight TSG column cells w/ 0 attendance when student is in initiative
def highlight_tsg_no_attendance(worksheet):
    ini_col = cld_col_num['Initiatives']
    exit_date_num = cld_col_num['Enrollment Exit Date']
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=ini_col, max_col=ini_col):
        for cell in row:
            exit_status = worksheet[get_column_letter(exit_date_num) + str(cell.row)].value
            if exit_status == datetime.datetime(9999, 12, 31, 0, 0):
                for i in initiatives:
                    if cell.value is not None:
                        tsg_list = list(str(cell.value).split('; '))
                    else:
                        tsg_list = []
                    if i in tsg_list and i != 'Family Advocacy Committee':
                        col_offset = cld_col_num[i] - ini_col #Distance between specific TSG column and 'Initiatives' column
                        if cell.offset(column=col_offset).value == 0:
                            pattern_fill(cell.offset(column=col_offset), color=Color.YELLOW)                    
                    