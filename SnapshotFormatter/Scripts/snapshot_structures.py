from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from datetime import datetime, date

workbook = load_workbook('Snapshots\\All School Snapshot Prep Flow Output.xlsx')
cld_ws = workbook['Caseload Details']
tr1_ws = workbook['Tier 1']
cie_ws = workbook['Check-ins']

#Create a dictionary of column names for caseload details. Note: openpyxl starts col_index at 1
cld_col_num = {}
current  = 1
for col in cld_ws.iter_cols(1, cld_ws.max_column):
    cld_col_num[col[0].value] = current
    current += 1

#Create a dictionary of column names for Tier 1. Note: openpyxl starts col_index at 1
tr1_col_num = {}
current  = 1
for col in tr1_ws.iter_cols(1, tr1_ws.max_column):
    tr1_col_num[col[0].value] = current
    current += 1    
    
#Create a letter-to-number dictionary
let_2_num = {chr(i):(i-64) for i in range(65, 91)}
    
#Set today's date
today_str = date.today().strftime("%m-%d-%y")
today = datetime.today()

#Create a dictionary of column names for Check-ins. Note: openpyxl starts col_index at 1
cie_col_num = {}
current  = 1
for col in cie_ws.iter_cols(1, cie_ws.max_column):
    cie_col_num[col[0].value] = current
    current += 1
    
#Create TSG initiatives list
    #Order matters
initiatives = ['SEL Lunch Group',
               'Snacktime',
               'Z Club',
              ]

#Create School-to-TSG dictionary
#Update regularly, TSG order should match the initiatives list
school_tsg = {'ABC Elementary': ['SEL Lunch Group', 'Snacktime'],
              '123 Middle School': ['SEL Lunch Group', 'Snacktime', 'Z Club'],
              'Generic High School': ['SEL Lunch Group', 'Z Club'],
              'All School': []
             }

#Create tier 2 activity keywords
tier2_kw = [' sel ',
            '(sel',
            'sel:',
            '1st ladies',
            'akemi',
            'angels',
            'bgm',
            'black girl magic',
            'black men',
            'black student support',
            'bois',
            "boy's group",
            'boys group',
            'bsap',
            'bsg',
            'bsu',
            'bushido',
            'estudiantiles',
            'first ladies',
            'game club',
            'gentlemen\'s group',
            'gentlemens group',
            'girl\'s group',
            'girlie pop',
            'girls group',
            'girls are magic',
            'golden girls',
            'herstory',
            'international students',
            'jones',
            'ladies 1st',
            'ladies first',
            'lamb',
            'lgbt',
            'lgbtqia',
            'little ladies',
            'megaladons',
            'men of action',
            'men of markham',
            'mixed gender',
            'newcomer',
            'prism',
            'regardless of color',
            'ronin',
            'salorzano',
            'sander',
            'scholarz',
            'sel group',
            'sony',
            'the unknowns',
            'tsg',
            'warrior',
            'wellness',
            'xy'
           ]


# set order for months
month_order = ['August', 'September', 'October', 'November', 'December', 'January', 'February', 'March', 'April', 'May', 'June']

# get current SY years
if today.month < 7:
    years = [today.year - 1, today.year]
else:
    years = [today.year, today.year + 1]

# set the last day of feb for the current SY
if years[1] % 4 == 0:
    end_feb = 29
else:
    end_feb = 28
    
#Create Month Dictionaries for the Check-ins and SC Entries tabs
months_end = {'August': datetime(years[0], 8, 31, 0, 0),
            'September': datetime(years[0], 9, 30, 0, 0),
            'October': datetime(years[0], 10, 31, 0, 0),
            'November': datetime(years[0], 11, 30, 0, 0),
            'December': datetime(years[0], 12, 31, 0, 0),
            'January': datetime(years[1], 1, 31, 0, 0),
            'February': datetime(years[1], 2, end_feb, 0, 0),
            'March': datetime(years[1], 3, 31, 0, 0),
            'April': datetime(years[1], 4, 30, 0, 0),
            'May': datetime(years[1], 5, 31, 0, 0),
            'June': datetime(years[1], 6, 30, 0, 0),
         }

months_beg = {'August': datetime(years[0], 8, 1, 0, 0),
            'September': datetime(years[0], 9, 1, 0, 0),
            'October': datetime(years[0], 10, 1, 0, 0),
            'November': datetime(years[0], 11, 1, 0, 0),
            'December': datetime(years[0], 12, 1, 0, 0),
            'January': datetime(years[1], 1, 1, 0, 0),
            'February': datetime(years[1], 2, 1, 0, 0),
            'March': datetime(years[1], 3, 1, 0, 0),
            'April': datetime(years[1], 4, 1, 0, 0),
            'May': datetime(years[1], 5, 1, 0, 0),
            'June': datetime(years[1], 6, 1, 0, 0),
         }

# Create a dictionary for caselod target numbers
caseload_projection = {'ABC Elementary': 50,
                       '123 Middle School': 100,
                       'Generic High School': 50,
                       'All School': 200
                      }
