from openpyxl.styles import Alignment

from snapshot_structures import *
from colors import Color, pattern_fill

# Formats each column to be the width of its widest cell and make header blue
def format_header(book):
    # Format Caseload Details sheet column dimensions
    cld = book['Caseload Details']
    cld.row_dimensions[1].height = 30
    for row in cld.iter_rows(min_row=1, max_row=1, min_col=1, max_col=cld.max_column):
        for cell in row:      
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)   
    for column_cells in cld.columns:
        length = max(len(str(cell.value)) for cell in column_cells)/2 + 7
        cld.column_dimensions[column_cells[0].column_letter].width = length
    
    # Format all other sheets column dimensions
    for sheet in book:
        for column_cells in sheet.columns:
            if sheet.title != 'Caseload Details':
                length = max(len(str(cell.value)) for cell in column_cells) + 2
                sheet.column_dimensions[column_cells[0].column_letter].width = length
            
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=sheet.max_column):
            for cell in row:
                pattern_fill(cell, color=Color.BLUE)
        
        #Format Notes columns
        notes1 = False # for tier2 sheet
        notes2 = False # for tier1 sheet
        if sheet.max_column > 6:
            if sheet['H1'].value == 'Notes':
                notes1 = True
            if sheet['N1'].value == 'Notes':
                notes2 = True
                
        if notes1 == True:
            sheet.column_dimensions['H'].width = 100
            for row in sheet.iter_rows(min_col=8, max_col=8):
                for cell in row:      
                    cell.alignment = Alignment(wrap_text=True,vertical='top')
                    
        if notes2 == True:
            sheet.column_dimensions['N'].width = 100
            for row in sheet.iter_rows(min_col=14, max_col=14):
                for cell in row:      
                    cell.alignment = Alignment(wrap_text=True,vertical='top')
